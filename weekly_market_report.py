"""
weekly_market_report.py
------------------------
Builds a weekly market snapshot into ONE master Excel workbook, adding a NEW
dated sheet each Monday. Designed to run Monday ~4:00 PM Dallas time (CT), by
which point that day's equity closes and NYMEX settlements are final.

Each weekly sheet, top to bottom:
  1. Quotes  — indices, EAFE proxy (EFA), FX pairs, equities — with pull time
  2. Energy — Henry Hub natural gas monthly strip (front month → Dec 2028) in a
              vertical column, with WTI crude in a matching column to its right
  3. Treasury par yield curve (kept from earlier; say the word to drop it)

Data sources (all free, no login):
  - Yahoo Finance          → indices, ETFs, FX, energy futures
  - U.S. Treasury XML feed  → yield curve

SETUP:  pip install yfinance openpyxl requests
"""

import yfinance as yf
import openpyxl
import requests
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, datetime, timedelta
from xml.etree import ElementTree as ET
import os
import sys
import time

try:
    from zoneinfo import ZoneInfo          # Python 3.9+
    CENTRAL = ZoneInfo("America/Chicago")
except Exception:                          # fallback if tzdata missing
    CENTRAL = None

# ══════════════════════════════════════════════════════════════════════════════
#  CONFIGURATION  —  EDIT THIS SECTION
# ══════════════════════════════════════════════════════════════════════════════

# Output path. Priority: command-line arg  >  MARKET_REPORT_PATH env var  >  default.
# Default writes to the current working directory (handy when run as a skill).
def _resolve_output_path():
    if len(sys.argv) > 1 and sys.argv[1].strip():
        return sys.argv[1].strip()
    env = os.environ.get("MARKET_REPORT_PATH", "").strip()
    if env:
        return env
    return os.path.abspath("Weekly_Market_Report.xlsx")

EXCEL_FILE = _resolve_output_path()

# Individual equities (Yahoo symbols)
STOCK_TICKERS = ["RPRX"]                    # Royalty Pharma

# Indices (Yahoo uses ^ prefixes)
INDEX_TICKERS = {
    "^DJI":  "Dow Jones Industrial Avg (DJIA)",
    "^GSPC": "S&P 500 (SPX)",
}

# International developed markets — EAFE via the iShares ETF proxy (EFA).
# The MSCI EAFE index level itself is not freely available; EFA tracks it.
EAFE_PROXY = {"EFA": "iShares MSCI EAFE ETF (EAFE proxy)"}

# FX pairs (Yahoo uses the =X suffix)
FX_TICKERS = {
    "EURUSD=X": "Euro / US Dollar (EURUSD)",
    "GBPUSD=X": "British Pound / US Dollar (GBPUSD)",
    "USDJPY=X": "US Dollar / Japanese Yen (USDJPY)",
}

# Energy strip runs from the front (non-expired) month through this year-end.
ENERGY_STRIP_END_YEAR = 2028

INCLUDE_TREASURY = True

# ══════════════════════════════════════════════════════════════════════════════
#  STYLING
# ══════════════════════════════════════════════════════════════════════════════

TITLE_FONT   = Font(name="Arial", bold=True, size=14, color="1F4E79")
STAMP_FONT   = Font(name="Arial", size=9,  italic=True, color="606B7A")
SECTION_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFF")
SECTION_FILL = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HEADER_FONT  = Font(name="Arial", bold=True, size=10, color="FFFFFF")
HEADER_FILL  = PatternFill("solid", start_color="4472C4", end_color="4472C4")
LABEL_FONT   = Font(name="Arial", bold=True, size=10)
DATA_FONT    = Font(name="Arial", size=10)
NA_FONT      = Font(name="Arial", size=10, italic=True, color="999999")

THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
RIGHT  = Alignment(horizontal="right")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left")

# ══════════════════════════════════════════════════════════════════════════════
#  FUTURES SYMBOLS
# ══════════════════════════════════════════════════════════════════════════════

MONTH_CODES = {1:"F",2:"G",3:"H",4:"J",5:"K",6:"M",7:"N",8:"Q",9:"U",10:"V",11:"X",12:"Z"}
MONTH_NAMES = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",
               7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}


def contract_label(year, month):
    return f"{MONTH_NAMES[month]}-{str(year)[-2:]}"


def energy_symbols(root, year, month):
    """Return the two Yahoo symbol formats to try for a NYMEX monthly contract.
    Near months usually resolve via '=F'; deferred months via the '.NYM' suffix."""
    code = MONTH_CODES[month]
    yy = str(year)[-2:]
    return [f"{root}{code}{yy}=F", f"{root}{code}{yy}.NYM"]


def energy_months(as_of):
    """Every (year, month) from the month AFTER as_of through ENERGY_STRIP_END_YEAR.
    The current delivery month has already expired by the time we run, so we start
    at the next month; fully-empty leading rows are trimmed after fetching."""
    months = []
    y, m = as_of.year, as_of.month
    # advance to next month
    m += 1
    if m > 12:
        m = 1; y += 1
    while y < ENERGY_STRIP_END_YEAR or (y == ENERGY_STRIP_END_YEAR and m <= 12):
        months.append((y, m))
        m += 1
        if m > 12:
            m = 1; y += 1
    return months

# ══════════════════════════════════════════════════════════════════════════════
#  DATA FETCHING
# ══════════════════════════════════════════════════════════════════════════════

# Yahoo throttles bursts of requests from cloud servers with HTTP 429. To stay
# under that, we (a) fetch in as few grouped calls as possible, (b) go sequential
# rather than parallel, (c) pause between call groups, and (d) wait-and-retry with
# growing delays whenever a batch comes back completely empty (the tell-tale sign
# of a rate-limit rather than a genuinely missing contract).

RETRY_DELAYS = [5, 15, 30, 60]   # seconds between retries of an empty batch
GROUP_PAUSE  = 3                 # polite pause between call groups

# Friendly names for equities (kept static to avoid extra rate-limited .info calls)
EQUITY_NAMES = {"RPRX": "Royalty Pharma plc"}


def _extract_close(data, sym, single):
    """Pull the latest close for one symbol out of a yf.download result."""
    try:
        col = (data["Close"] if single else data[sym]["Close"]).dropna()
        if len(col):
            return round(float(col.iloc[-1]), 4)
    except Exception:
        pass
    return None


def batched_closes(symbols, label=""):
    """Fetch the latest close for a list of symbols in one grouped call, waiting
    and retrying if Yahoo rate-limits us. Returns {symbol: price_or_None}."""
    result = {s: None for s in symbols}
    if not symbols:
        return result

    for attempt in range(len(RETRY_DELAYS) + 1):
        if attempt > 0:
            delay = RETRY_DELAYS[attempt - 1]
            print(f"    (Yahoo throttled {label}; waiting {delay}s then retrying "
                  f"[{attempt}/{len(RETRY_DELAYS)}])")
            time.sleep(delay)
        try:
            data = yf.download(symbols, period="5d", progress=False,
                               group_by="ticker", threads=False)
        except Exception as e:
            print(f"    (download error for {label}: {e})")
            data = None

        got_any = False
        if data is not None and not getattr(data, "empty", True):
            single = (len(symbols) == 1)
            for s in symbols:
                if result[s] is None:
                    px = _extract_close(data, s, single)
                    if px is not None:
                        result[s] = px
                        got_any = True

        if all(v is not None for v in result.values()):
            break            # everything filled — done
        if got_any:
            break            # partial success → missing ones are likely just
                             # unavailable contracts, not throttling; stop retrying
        # else: got nothing at all → probably throttled → loop and retry

    return result


def fetch_quotes():
    """Indices, EAFE proxy, FX, equities → list of (sym, name, price, group).
    All fetched in ONE grouped call to minimize requests."""
    print("  Quotes: fetching indices, EAFE, FX, equities in one batch...")
    spec = []  # (sym, name, group)
    for sym, name in INDEX_TICKERS.items():
        spec.append((sym, name, "Index"))
    for sym, name in EAFE_PROXY.items():
        spec.append((sym, name, "Intl"))
    for sym, name in FX_TICKERS.items():
        spec.append((sym, name, "FX"))
    for sym in STOCK_TICKERS:
        spec.append((sym, EQUITY_NAMES.get(sym, sym), "Equity"))

    prices = batched_closes([s for s, _, _ in spec], label="quotes")
    rows = [(sym, name, prices.get(sym), grp) for (sym, name, grp) in spec]
    for sym, name, price, grp in rows:
        print(f"    {sym:<10} {price if price is not None else 'N/A'}")
    return rows


def fetch_energy_curve(root, months):
    """Fetch settlement/last for each (year,month) contract of `root`.
    Tries the '=F' format in one batch, then '.NYM' for anything still missing.
    Returns {(year,month): price_or_None}."""
    prices = {ym: None for ym in months}

    # Pass 1: '=F' symbols, one grouped call (with throttle-retry)
    map_f = {energy_symbols(root, y, m)[0]: (y, m) for (y, m) in months}
    got_f = batched_closes(list(map_f.keys()), label=f"{root} strip")
    for sym, ym in map_f.items():
        if got_f.get(sym) is not None:
            prices[ym] = got_f[sym]

    # Pass 2: '.NYM' only for the still-missing months
    missing = [ym for ym in months if prices[ym] is None]
    if missing:
        time.sleep(GROUP_PAUSE)
        map_nym = {energy_symbols(root, y, m)[1]: (y, m) for (y, m) in missing}
        got_n = batched_closes(list(map_nym.keys()), label=f"{root} deferred")
        for sym, ym in map_nym.items():
            if got_n.get(sym) is not None:
                prices[ym] = got_n[sym]

    got = sum(1 for v in prices.values() if v is not None)
    print(f"    {root}: {got}/{len(months)} contracts returned a quote")
    return prices


# ── Treasury feed ─────────────────────────────────────────────────────────────
TREASURY_URL = ("https://home.treasury.gov/resource-center/data-chart-center/"
                "interest-rates/pages/xml")
NS   = "http://schemas.microsoft.com/ado/2007/08/dataservices"
NS_M = "http://schemas.microsoft.com/ado/2007/08/dataservices/metadata"
TREASURY_FIELDS = [
    ("BC_1MONTH","1 Mo"),("BC_2MONTH","2 Mo"),("BC_3MONTH","3 Mo"),
    ("BC_4MONTH","4 Mo"),("BC_6MONTH","6 Mo"),("BC_1YEAR","1 Yr"),
    ("BC_2YEAR","2 Yr"),("BC_3YEAR","3 Yr"),("BC_5YEAR","5 Yr"),
    ("BC_7YEAR","7 Yr"),("BC_10YEAR","10 Yr"),("BC_20YEAR","20 Yr"),
    ("BC_30YEAR","30 Yr"),
]

def fetch_treasury_yields(target):
    print("  Treasury yield curve...")
    def pull_month(year, month):
        params = {"data": "daily_treasury_yield_curve",
                  "field_tdr_date_value_month": f"{year}{month:02d}"}
        r = requests.get(TREASURY_URL, params=params, timeout=30)
        r.raise_for_status()
        root = ET.fromstring(r.content)
        recs = []
        for props in root.findall(f".//{{{NS_M}}}properties"):
            d = props.find(f"{{{NS}}}NEW_DATE")
            if d is None or not d.text:
                continue
            rec = {"date": d.text[:10]}
            for field, _ in TREASURY_FIELDS:
                el = props.find(f"{{{NS}}}{field}")
                rec[field] = float(el.text) if (el is not None and el.text) else None
            recs.append(rec)
        return recs
    try:
        records = pull_month(target.year, target.month)
        if not records:
            prev = target.replace(day=1) - timedelta(days=1)
            records = pull_month(prev.year, prev.month)
        target_str = target.strftime("%Y-%m-%d")
        match = next((r for r in records if r["date"] == target_str), None)
        if not match:
            elig = [r for r in records if r["date"] <= target_str]
            match = sorted(elig, key=lambda x: x["date"])[-1] if elig else None
        if not match:
            print("    (No Treasury data near target date.)")
            return None, {}
        return match["date"], {lbl: match[f] for f, lbl in TREASURY_FIELDS}
    except Exception as e:
        print(f"    (Treasury fetch failed: {e} — section skipped.)")
        return None, {}

# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL WRITING
# ══════════════════════════════════════════════════════════════════════════════

def put(ws, r, c, val, font=DATA_FONT, align=RIGHT, num_fmt=None, border=True):
    cell = ws.cell(r, c, val)
    cell.font = font
    cell.alignment = align
    if num_fmt:
        cell.number_format = num_fmt
    if border:
        cell.border = BORDER
    return cell

def section_bar(ws, r, text, span):
    c = ws.cell(r, 1, text)
    c.font = SECTION_FONT
    for col in range(1, span + 1):
        ws.cell(r, col).fill = SECTION_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=span)
    ws.row_dimensions[r].height = 20
    return r + 1

def hcell(ws, r, c, text):
    cell = put(ws, r, c, text, font=HEADER_FONT, align=CENTER)
    cell.fill = HEADER_FILL
    return cell


def build_weekly_sheet(wb, sheet_name, as_of_str, pulled_str,
                       quotes, ng_prices, cl_prices, energy_months_list,
                       tsy_date, tsy_yields):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    # Title + pull stamp
    t = ws.cell(1, 1, f"Weekly Market Data — Week of {as_of_str}")
    t.font = TITLE_FONT
    ws.cell(2, 1,
            f"Data as of {as_of_str} market close (Dallas / Central Time)   •   "
            f"Pulled {pulled_str}").font = STAMP_FONT

    row = 4

    # ── Section 1: Quotes ─────────────────────────────────────────────────────
    row = section_bar(ws, row, "  QUOTES — INDICES, EAFE, FX & EQUITIES  (closing level)", 4)
    for col, h in enumerate(["Symbol", "Name", "Type", "Close"], start=1):
        hcell(ws, row, col, h)
    row += 1
    for (sym, name, price, group) in quotes:
        put(ws, row, 1, sym, font=LABEL_FONT, align=LEFT)
        put(ws, row, 2, name, align=LEFT)
        put(ws, row, 3, group, align=CENTER)
        if price is None:
            put(ws, row, 4, "N/A", font=NA_FONT)
        else:
            fmt = "#,##0.0000" if group == "FX" else "#,##0.00"
            put(ws, row, 4, price, num_fmt=fmt)
        row += 1
    row += 2

    # ── Section 2: Energy (vertical NG column, WTI column to the right) ────────
    # Columns: A=NG month, B=NG price, C=spacer, D=WTI month, E=WTI price
    row = section_bar(
        ws, row,
        f"  ENERGY — MONTHLY SETTLEMENTS AS OF {as_of_str}  (front month → Dec {ENERGY_STRIP_END_YEAR})",
        5)
    hcell(ws, row, 1, "HH Nat Gas Contract")
    hcell(ws, row, 2, "$/MMBtu")
    ws.cell(row, 3, None)
    hcell(ws, row, 4, "WTI Crude Contract")
    hcell(ws, row, 5, "$/bbl")
    row += 1

    for (y, m) in energy_months_list:
        label = contract_label(y, m)
        ng = ng_prices.get((y, m))
        cl = cl_prices.get((y, m))
        # NG block
        put(ws, row, 1, label, font=LABEL_FONT, align=LEFT)
        if ng is None:
            put(ws, row, 2, "—", font=NA_FONT, align=CENTER)
        else:
            put(ws, row, 2, ng, num_fmt="#,##0.000")
        # spacer (no border)
        ws.cell(row, 3, None)
        # WTI block
        put(ws, row, 4, label, font=LABEL_FONT, align=LEFT)
        if cl is None:
            put(ws, row, 5, "—", font=NA_FONT, align=CENTER)
        else:
            put(ws, row, 5, cl, num_fmt="#,##0.000")
        row += 1
    row += 2

    # ── Section 3: Treasury ───────────────────────────────────────────────────
    if tsy_yields:
        labels = [lbl for _, lbl in TREASURY_FIELDS]
        row = section_bar(ws, row,
                          f"  TREASURY PAR YIELD CURVE  (as of {tsy_date})  %",
                          len(labels) + 1)
        hcell(ws, row, 1, "Maturity")
        for i, lbl in enumerate(labels, start=2):
            hcell(ws, row, i, lbl)
        row += 1
        put(ws, row, 1, "Yield %", font=LABEL_FONT, align=LEFT)
        for i, lbl in enumerate(labels, start=2):
            v = tsy_yields.get(lbl)
            if v is None:
                put(ws, row, i, "—", font=NA_FONT, align=CENTER)
            else:
                put(ws, row, i, v, num_fmt="0.00")
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 3
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 12
    for col in range(6, 16):
        ws.column_dimensions[get_column_letter(col)].width = 8
    return ws


def load_or_create_workbook(path):
    if os.path.exists(path):
        return openpyxl.load_workbook(path)
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    print(f"  Created new workbook: {path}")
    return wb

# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

def get_target_monday():
    today = date.today()
    return today - timedelta(days=today.weekday())

def pulled_stamp():
    now = datetime.now(CENTRAL) if CENTRAL else datetime.now()
    return now.strftime("%Y-%m-%d %H:%M %Z").strip()


def trim_leading_blank(months, ng_prices, cl_prices):
    """Drop leading months where BOTH products have no quote."""
    for i, ym in enumerate(months):
        if ng_prices.get(ym) is not None or cl_prices.get(ym) is not None:
            return months[i:]
    return months


def main():
    print("=== Weekly Market Report Builder ===\n")
    target = get_target_monday()
    as_of_str = target.strftime("%Y-%m-%d")
    pulled = pulled_stamp()
    sheet_name = as_of_str
    print(f"  Week of : {as_of_str}")
    print(f"  Pulled  : {pulled}\n")

    wb = load_or_create_workbook(EXCEL_FILE)
    if sheet_name in wb.sheetnames:
        print(f"  Sheet '{sheet_name}' already exists — nothing to do.")
        return

    quotes = fetch_quotes()
    time.sleep(GROUP_PAUSE)

    months = energy_months(target)
    print(f"\n  Energy strip: {contract_label(*months[0])} → "
          f"{contract_label(*months[-1])} ({len(months)} months)")
    ng_prices = fetch_energy_curve("NG", months)
    time.sleep(GROUP_PAUSE)
    cl_prices = fetch_energy_curve("CL", months)
    months = trim_leading_blank(months, ng_prices, cl_prices)

    if INCLUDE_TREASURY:
        tsy_date, tsy_yields = fetch_treasury_yields(target)
    else:
        tsy_date, tsy_yields = None, {}

    build_weekly_sheet(wb, sheet_name, as_of_str, pulled, quotes,
                       ng_prices, cl_prices, months, tsy_date, tsy_yields)
    wb.move_sheet(sheet_name, -(len(wb.sheetnames) - 1))
    wb.save(EXCEL_FILE)
    print(f"\n  Added sheet '{sheet_name}' → {EXCEL_FILE}")
    print("Done.")


if __name__ == "__main__":
    main()
